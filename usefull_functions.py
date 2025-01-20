import os
import openpyxl
import settings
import pages.page_errorschecker
from openpyxl.styles import PatternFill
from subject_class import Subject

def files_list(directory):
    return os.listdir(directory)


def get_schedule():
    wb = openpyxl.load_workbook(settings.current_file, read_only=False, keep_vba=True)

    db = []
    ws = wb.active
    j = 3
    for cols in ws.iter_cols(min_col=3, max_col=3, max_row=107):
        i = 4
        for cell in cols:
            if cell.value is not None or "/" in str(cell.value):
                db.append(Subject(
                    week_type=ws.cell(2, 1).value,
                    day_per_week=ws.cell(i, 1).value,
                    lesson_number=ws.cell(i, 2).value,
                    cabinet=ws.cell(i, j + 1).value,
                    teacher="",
                    group=ws.cell(2, j).value
                    )
                ),
                print(f"Значение i = {i}, j = {j}")
                print(f"Кабинет - {db[0].cabinet}")
                print(i, cell.value, cell.coordinate, end="\n\n")
                i += 1
        j += 1


def check_errors():
    wb = openpyxl.load_workbook(settings.current_file, read_only=False, keep_vba=True)

    ws = wb.active
    for rows in ws.iter_rows():
            for cell in rows:
                if cell.value is not None and "/" in str(cell.value):
                    cell.fill = PatternFill(patternType="solid", fgColor="00CC00")


get_schedule()